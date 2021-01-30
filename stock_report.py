import requests
import json
import pandas as pd
from pandas import Series, DataFrame
import numpy as np
import openpyxl
import os
def stock_report( name, code ):
    # =====================================================입력창===========================================
    API_KEY = "9603774f25edc5686b1d7df2e3d1f8788a864fe3"
    PARAMS = {
        'crtfc_key': API_KEY,
        # !!!!! 여기에 "고유코드"를 입력해주세요
        'corp_code': code,
        # 사업연도(4자리)
        'bsns_year': '2019',
        # 사업보고서
        # 1분기보고서 : 11013
        # 반기보고서 : 11012
        # 3분기보고서 : 11014
        # 사업보고서 : 11011
        'reprt_code': '11011',
    }
    # 엑셀에 입력될 회사 이름으로 바꿔주세요
    company_name = name
    # =====================================================입력창===========================================

    # Setting
    # 최대 줄 수 설정
    pd.set_option('display.max_rows', 500)
    # 최대 열 수 설정
    pd.set_option('display.max_columns', 500)
    # 표시할 가로의 길이
    pd.set_option('display.width', 1000)


    URL = 'https://opendart.fss.or.kr/api/fnlttSinglAcnt.json'


    resp = requests.get(url=URL, params=PARAMS)

    # http 정상응답시 처리
    if resp.status_code == 200:
        data_json = resp.json()

        # OUTPUT
        data_str = json.dumps(data_json, indent=4, ensure_ascii=False)
        # print(data_str)


        if data_json['status'] == "000":
            detail = data_json['list']

            # for x in detail:
            # if x['fs_div'] == 'CFS' and x['sj_div'] == 'IS' and x['account_nm'] == '당기순이익':
            # print(json.dumps(x, indent=4, ensure_ascii=False))

            # Json 코드 DataFrame으로 변환
            df = pd.json_normalize(detail)
            # print(df)
        else:
            pass
            # print(data_json['message'])



    # 특정컬럼 추출
    # _account_nm_sr = df['account_nm']
    # print('_account_nm_sr: ', type(_account_nm_sr))
    # print(_account_nm_sr)
    # N개 컬럼 추출
    _main_columns = df[['account_nm', 'fs_div', 'sj_div']]
    # print('_main_columns: ', type(_main_columns))
    # print(_main_columns)


    # 1. 조건을 Series로 선언한 후 조립
    _fs = df['fs_div'] == 'CFS'
    _sj = df['sj_div'] == 'IS'
    result_01 = df[_fs & _sj]
    # # 2. 내부에 바로 선언, 각 조건은 ()을 묶어줘야한다.
    # result_02 = df[
    #     (df['fs_div'] == 'CFS')
    #     & (df['sj_div'] == 'IS')
    #     ]
    # ** 내맘대로 조건 2로 재무생태표를 출력하게 함.
    _fs = df['fs_div'] == 'CFS'
    _sj = df['sj_div'] == 'BS'
    result_02 = df[_fs & _sj]

    # 계정과목명, 당기금액, 전기금액 추출
    _extract_cols = ['account_nm', 'bfefrmtrm_amount', 'frmtrm_amount', 'thstrm_amount']
    df_j = result_02.loc[:, _extract_cols]
    # print('==========재무상태표==========')
    # print(df_j)
    # 계정과목명, 당기금액, 전기금액 추출
    _extract_cols = ['account_nm', 'bfefrmtrm_amount', 'frmtrm_amount', 'thstrm_amount']
    df_s = result_01.loc[:, _extract_cols]
    # print('==========손익계산서==========')
    # print(df_s)



    # comma(,) 제거 후 np.int64 파싱
    df_s.loc[:, ['bfefrmtrm_amount', 'frmtrm_amount', 'thstrm_amount']] = df_s[['bfefrmtrm_amount', 'frmtrm_amount', 'thstrm_amount']].apply(
        lambda x: x.str.replace(',', '').astype(np.int64))
    df_j.loc[:, ['bfefrmtrm_amount', 'frmtrm_amount', 'thstrm_amount']] = df_j[['bfefrmtrm_amount', 'frmtrm_amount', 'thstrm_amount']].apply(
        lambda x: x.str.replace(',', '').astype(np.int64))

    # extracted_df.loc[:, 'ratio'] = extracted_df.thstrm_amount / extracted_df.frmtrm_amount * 100
    #
    # extracted_df.loc[:, 'gap'] = extracted_df.thstrm_amount - extracted_df.frmtrm_amount
    #
    # print(extracted_df)

    # **** dataframe to list
    # list = np.array(extracted_df_j[1].tolist())
    # print(list)

    #=======================================================================================

    sgod_list = []
    sgod_list.append(company_name)
    per_m_list = []
    per_y_list = []
    def make_sgod_list():
        for i in range(0,4):
            if i != 2:
                df_tolist = list(np.array(df_s.iloc[i].tolist()))
                del df_tolist[0]
                for j in range (0,3):
                    change_num = int(df_tolist[j])

                    if i==0:
                        per_m_list.append(change_num)
                    if i==1:
                        per_y_list.append(change_num)

                    if change_num < 100000000000:
                        list_new = round( change_num /100000000, 1)
                    else:
                        list_new = round( change_num /100000000 )
                    list_new = make_comma(list_new)
                    sgod_list.append(list_new)
            if i == 2:
                for z in range(0,3):
                    this_one = round( per_y_list[z]/per_m_list[z]*100, 1)
                    this_one = str(this_one)+"%"
                    sgod_list.append(this_one)

    # print(god_list)

    # # call xlsx
    # # wb = openpyxl.Workbook()
    # wb = openpyxl.load_workbook("df_s.xlsx")
    #
    # sheet = wb.active
    # sheet.append(god_list)
    # wb.save("df_s.xlsx")

    jgod_list = []
    jgod_list.append(company_name)
    def make_jgod_list():
        # df_yj = list(np.array(df_s.iloc[0].tolist()))
        # df_yb = list(np.array(df_s.iloc[3].tolist()))
        # df_jc = list(np.array(df_s.iloc[2].tolist()))
        # df_bc = list(np.array(df_s.iloc[5].tolist()))
        # df_jj = list(np.array(df_s.iloc[8].tolist()))
        # df_jg = list(np.array(df_s.iloc[6].tolist()))

        navi1 = [0,5,3,2,5,8,8]
        navi2 = [3,8,0,0,0,2,6]
        for m in range(0,7):
            if m==2 or m==3 or m==4:
                df_tolist = list(np.array(df_j.iloc[ navi1[m] ].tolist()))
                jgod_list.append( katarina(df_tolist[1], df_tolist[2], 1))
                jgod_list.append( katarina(df_tolist[2], df_tolist[3], 1))
            else:
                df_tolist1 = list(np.array(df_j.iloc[ navi1[m] ].tolist()))
                df_tolist2 = list(np.array(df_j.iloc[ navi2[m] ].tolist()))
                if m==6:
                    if katarina(df_tolist1[3], df_tolist2[3], 6):
                        jgod_list.append('!!!')
                    else:
                        jgod_list.append(' ')
                else:
                    jgod_list.append( katarina(df_tolist1[2], df_tolist2[2], 0) )
                    jgod_list.append( katarina(df_tolist1[3], df_tolist2[3], 0) )


    # a/b || a==전기 b==당기
    def katarina(a, b, c):
        if c == 0:
            this_one = round( int(a) / int(b) * 100, 1)
            return str(this_one)+'%'
        elif c == 1:
            a = int(a)
            b = int(b)
            b_a = b-a
            this_one = round( b_a/a * 100, 1)
            return str(this_one)+'%'
        elif c == 6:
            a = int(a)
            b = int(b)
            if a < b:
                return 1
            else:
                return 0

    def make_comma(num):
        return "{:,}".format(num)


    # call xlsx
    # wb = openpyxl.Workbook()
    def write_xlsx(name, god_list):
        wb = openpyxl.load_workbook(name)

        sheet = wb.active
        sheet.append(god_list)
        wb.save(name)


    make_sgod_list()
    make_jgod_list()

    write_xlsx("df_j.xlsx", jgod_list)
    write_xlsx("df_s.xlsx", sgod_list)



# print('==================================================================')
# print(df_s.iloc[0])
# print('==================================================================')
# df_test = list(np.array(df_s.iloc[0].tolist()))
# del df_test[0]
# print(df_test)


#=======================================================================================

# # *** to Excel
# base_dir = "C:/Users/HR/double/stock"
# file_nm = "df.xlsx"
# xlsx_dir = os.path.join(base_dir, file_nm)
#
# new_df = extracted_df_j
#
# new_df.to_excel(xlsx_dir,
#                 sheet_name = 'Sheet1',
#                 na_rep = 'NaN',
#                 float_format= "%.sf",
#                 header= True,
#                 index = True,
#                 index_label = "id" ,
#                 startrow=1,
#                 startcol=1,
#                 freeze_panes= (2,0)
#                 )
