# -*- coding: utf-8 -*-
from bs4 import BeautifulSoup
import requests
import openpyxl
from tqdm import tqdm

class Crawl:
    def __init__(self):
        self.c_indicator = ['p', 1, 2, 4, 3, 5, 10, 11, 6]
        self.issimple = 0
        self.line_list = []

    def update_price(self):
        excel = Excel()
        code_list = excel.update_price_excel()
        count = 3
        for code in tqdm(code_list):
            self.code = code
            self.get_html()
            if self.resp.status_code == 200:
                element = BeautifulSoup(self.html, 'html.parser').select_one('p.no_today').text.strip().split('\n')
                excel.save_a_price(element[0], count)
                count += 1
            else:
                print('ERROR ! in update_price, class Crawl')

    def copy_excel(self, exfilename, year):
        excel = Excel()
        code_list = excel.update_price_excel(exfilename)[0]
        name_list = excel.update_price_excel(exfilename)[1]
        print('code_list')
        print(f'{code_list}')
        print('name_list')
        print(f'{name_list}')
        for i in tqdm(range( len(code_list) )):
            self.update_single(name_list[i], code_list[i], year)

    def update_single(self, name, code, year):
        self.line_list = []
        self.year = year
        self.code = code
        self.get_html()
        self.name = name
        if self.resp.status_code == 200:
            self.soup = BeautifulSoup(self.html, 'html.parser')
            if self.issimple == 1:
                return self.price_check()
            else:
                self.year_check()
                self.crawl_it()

    def year_check(self):
        url = 'div.cop_analysis thead tr:nth-child(2) th:nth-child(4)'
        n_year = self.soup.select_one(url).text.strip()[:4]
        if n_year == self.year:
            self.year_matched = 1
        else:
            self.year_matched = 0

    def crawl_it(self):
        if self.year_matched == 1:
            n = 3
        else:
            n = 2
        for x in self.c_indicator:
            if x == 'p':
                url = 'p.no_today'
                self.insert_list(url)
            elif x == 'mp':
                pass
            else:
                if x == 10 or x == 11 or x == 6:
                    for i in range(n, n+2):
                        url = 'div.cop_analysis table tbody tr:nth-child(' + str(x) + ') td:nth-child(' + str(i) + ')'
                        self.insert_list(url)
                else:
                    for i in range(n, n+3):
                        url = 'div.cop_analysis table tbody tr:nth-child(' + str(x) + ') td:nth-child(' + str(i) + ')'
                        self.insert_list(url)
        self.remake_list()
        code_list = self.make_code_list()
        excel = Excel()
        excel.save_excel(self.line_list, code_list)

    def insert_list(self, url):
        element = self.soup.select_one(url).text.strip()
        if url == 'p.no_today':
            self.line_list.append(self.name)
            sosok = self.soup.select_one('#_market_sum').text.strip()+"억"
            element = element.split('\n')
            self.line_list.append(element[0])
            self.line_list.append(sosok)
        else:
            self.line_list.append(element)

    def remake_list(self):
        list = self.line_list
        mud_list = []
        for i in range(2):
            a = self.rc(list[i+4])
            b = self.rc(list[i+3])
            print(a)
            if a is not None:
                this_one = round( (a-b)/b*100, 1 )
            else:
                this_one = ''
            mud_list.append(this_one)
        mud_list.reverse()
        for item in mud_list:
            list.insert(6, item)
        self.line_list = list

    def rc(self, x):
        # remove_comma
        if ',' in x:
            return int(x.replace(',', ''))
        elif x == '':
            pass
        else:
            return int(x)

    def make_code_list(self):
        result = ['', self.code]
        return result

    def get_html(self):
        url = 'https://finance.naver.com/item/main.nhn?code='+self.code
        self.resp = requests.get(url)
        self.html = self.resp.text

    def price_check(self):
        url = 'p.no_today'
        element = self.soup.select_one(url).text.strip()
        if element:
            return True
        else:
            return False

    def simple_check(self, code):
        self.issimple = 1
        result = self.update_single('', code)
        return result


class Excel:
    def __init__(self):
        self.name = 'stock.xlsx'
        self.code_list = []
        self.name_list = []

    def save_excel(self, list, code):
        wb = openpyxl.load_workbook(self.name)

        sheet = wb.get_sheet_by_name('기업정보')
        sheet.append(list)
        wb.save(self.name)

        sheet = wb.get_sheet_by_name('for_code')
        sheet.append(code)
        wb.save(self.name)

    def update_price_excel(self, exfilename):
        if exfilename is not None:
            self.wb = openpyxl.load_workbook(exfilename+'.xlsx')
        else:
            self.wb = openpyxl.load_workbook(self.name)
        self.ws = self.wb.get_sheet_by_name('기업정보')
        self.take_code_list()
        if exfilename is not None:
            self.take_name_list()
            return [self.code_list, self.name_list]
        else:
            return self.code_list

    def take_name_list(self):
        for i in range(3, 999):
            element = self.ws['A'+str(i)].value
            if element is None:
                break
            else:
                self.name_list.append(element)

    def take_code_list(self):
        ws = self.wb.get_sheet_by_name('for_code')
        for i in range(3, 999):
            element = ws['B'+str(i)].value
            if element is None:
                break
            else:
                self.code_list.append(element)

    def save_a_price(self, element, count):
        self.ws['B'+str(count)] = element
        self.wb.save(self.name)
        return